"""Tests fuer manuelle Zeiterfassung, Spaltenkonfiguration und Aufwand-Parser."""

from __future__ import annotations

from collections.abc import Iterator
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from jira_timesheet.models.export_column import (
    COLUMN_DEFAULTS,
    ExportColumn,
    default_columns,
    parse_columns,
    pdf_column_widths,
)
from jira_timesheet.models.settings import (
    DEFAULT_CUSTOMERS,
    DEFAULT_MANUAL_COLOR,
    Settings,
    normalize_color,
)
from jira_timesheet.models.timesheet import Timesheet, TimesheetDay, WorklogEntry
from jira_timesheet.services.duration import format_hours, parse_hours
from jira_timesheet.services.excel_exporter import ExcelExporter
from jira_timesheet.services.manual_entry_service import ManualEntry, ManualEntryService
from jira_timesheet.services.timesheet_service import TimesheetService


@pytest.fixture
def service(tmp_path: Path) -> Iterator[ManualEntryService]:
    """Repository auf einer temporaeren DB - nie auf der echten Nutzerdatei."""
    svc = ManualEntryService(db_path=tmp_path / "manual.db")
    svc.connect()
    yield svc
    svc.close()


class TestDurationParser:
    """Der Aufwand-Parser muss die Schreibweisen aus dem Ticket-Alltag koennen."""

    @pytest.mark.parametrize(
        ("text", "expected"),
        [
            ("3h 30m", 3.5),
            ("5h", 5.0),
            ("15m", 0.25),
            ("3:30", 3.5),
            ("0:15", 0.25),
            ("3,5", 3.5),
            ("3.5", 3.5),
            ("2h30", 2.5),
            ("3 h", 3.0),
        ],
    )
    def test_accepted_forms(self, text: str, expected: float) -> None:
        assert parse_hours(text) == pytest.approx(expected)

    @pytest.mark.parametrize("text", ["", "   ", "abc", "-1", "drei Stunden"])
    def test_rejected_forms(self, text: str) -> None:
        assert parse_hours(text) is None

    def test_format_roundtrip(self) -> None:
        for value in (3.5, 5.0, 0.25, 10.0):
            assert parse_hours(format_hours(value)) == pytest.approx(value)


class TestManualEntryService:
    """SQLite-Ablage: anlegen, lesen, aendern, loeschen."""

    def test_add_and_read_back(self, service: ManualEntryService) -> None:
        entry_id = service.add(
            ManualEntry(
                entry_date=date(2026, 7, 1),
                ticket="DMZ-17024",
                summary="GitLab-Pipelines",
                customer="Corporate",
                hours=10.0,
            )
        )
        assert entry_id > 0

        stored = service.get(entry_id)
        assert stored is not None
        assert stored.ticket == "DMZ-17024"
        assert stored.customer == "Corporate"
        assert stored.hours == pytest.approx(10.0)
        assert stored.created_at

    def test_entries_between_filters_by_date(self, service: ManualEntryService) -> None:
        for day, hours in ((date(2026, 6, 30), 8.0), (date(2026, 7, 1), 10.0), (date(2026, 8, 1), 2.0)):
            service.add(ManualEntry(entry_date=day, ticket="DMZ-17024", hours=hours))

        july = service.entries_between(date(2026, 7, 1), date(2026, 7, 31))
        assert [e.entry_date for e in july] == [date(2026, 7, 1)]

    def test_worklogs_carry_manual_flag_and_id(self, service: ManualEntryService) -> None:
        entry_id = service.add(
            ManualEntry(entry_date=date(2026, 7, 1), ticket="DMZ-17024", hours=10.0, customer="Corporate")
        )
        worklogs = service.worklogs_between(date(2026, 7, 1), date(2026, 7, 31), author="Michael")
        assert len(worklogs) == 1
        assert worklogs[0].manual is True
        assert worklogs[0].manual_id == entry_id
        assert worklogs[0].customer == "Corporate"
        assert worklogs[0].author == "Michael"

    def test_update_changes_values(self, service: ManualEntryService) -> None:
        entry_id = service.add(ManualEntry(entry_date=date(2026, 7, 1), ticket="DMZ-1", hours=1.0))
        stored = service.get(entry_id)
        assert stored is not None
        stored.hours = 3.5
        stored.ticket = "DMZ-17024"
        assert service.update(stored) is True

        again = service.get(entry_id)
        assert again is not None
        assert again.hours == pytest.approx(3.5)
        assert again.ticket == "DMZ-17024"

    def test_delete_removes_row(self, service: ManualEntryService) -> None:
        entry_id = service.add(ManualEntry(entry_date=date(2026, 7, 1), ticket="DMZ-1", hours=1.0))
        assert service.delete(entry_id) is True
        assert service.get(entry_id) is None
        assert service.delete(entry_id) is False

    def test_migrate_is_idempotent(self, tmp_path: Path) -> None:
        """Zweiter Start auf derselben Datei darf nicht scheitern."""
        db = tmp_path / "manual.db"
        with ManualEntryService(db_path=db) as first:
            first.add(ManualEntry(entry_date=date(2026, 7, 1), ticket="DMZ-1", hours=1.0))
        with ManualEntryService(db_path=db) as second:
            assert second.count() == 1


class TestMergeIntoTimesheet:
    """Manuelle Zeiten muessen in Tagessummen und Luecken-Erkennung eingehen."""

    def test_manual_hours_count_into_day_total(self) -> None:
        jira = WorklogEntry(
            date=date(2026, 7, 1), ticket="DMZ-13983", summary="Agile", author="M", budget="", hours=1.0
        )
        manual = ManualEntry(entry_date=date(2026, 7, 1), ticket="DMZ-17024", hours=10.0).to_worklog()

        sheet = TimesheetService.build_timesheet(
            entries=[jira, manual],
            developer="M",
            email="m@example.org",
            date_from=date(2026, 7, 1),
            date_to=date(2026, 7, 31),
        )
        assert sheet.total_hours == pytest.approx(11.0)
        assert sheet.days[0].total_hours == pytest.approx(11.0)

    def test_day_with_only_manual_hours_is_not_a_gap(self) -> None:
        """Der Merge muss vor der Lueckenerkennung passieren."""
        manual = ManualEntry(entry_date=date(2026, 7, 1), ticket="DMZ-17024", hours=10.0).to_worklog()
        worked_dates = {e.date for e in [manual]}
        assert date(2026, 7, 1) in worked_dates


class TestColumnConfiguration:
    """Spaltenkonfiguration: Defaults, Merge und Breitenverteilung."""

    def test_defaults_cover_all_columns(self) -> None:
        columns = default_columns()
        assert [c.key for c in columns] == [c.key for c in COLUMN_DEFAULTS]
        assert all(c.enabled for c in columns)
        assert all(c.label for c in columns)

    def test_parse_fills_missing_columns_from_defaults(self) -> None:
        """Eine alte settings.json ohne die Kunden-Spalte bekommt sie ergaenzt."""
        stored = [{"key": "week", "label": "Woche", "enabled": False}]
        columns = parse_columns(stored)
        assert len(columns) == len(COLUMN_DEFAULTS)
        assert columns[0].label == "Woche"
        assert columns[0].enabled is False
        customer = next(c for c in columns if c.key == "customer")
        assert customer.label == "Kunde"
        assert customer.enabled is True

    def test_parse_ignores_unknown_keys_and_garbage(self) -> None:
        columns = parse_columns([{"key": "nonexistent"}, "kaputt", 42])
        assert [c.key for c in columns] == [c.key for c in COLUMN_DEFAULTS]
        assert parse_columns(None) == default_columns()

    def test_empty_label_falls_back_to_default(self) -> None:
        columns = parse_columns([{"key": "ticket", "label": "   ", "enabled": True}])
        ticket = next(c for c in columns if c.key == "ticket")
        assert ticket.label == "Ticket"

    def test_pdf_widths_fill_the_page(self) -> None:
        widths = pdf_column_widths(default_columns())
        assert sum(widths) == pytest.approx(277.0)

    def test_pdf_widths_without_description_still_fill_the_page(self) -> None:
        columns = [c for c in default_columns() if c.key != "description"]
        widths = pdf_column_widths(columns)
        assert len(widths) == len(columns)
        assert sum(widths) == pytest.approx(277.0)


class TestCustomerList:
    """Kundenliste: Defaults, defensives Laden, bereits benutzte Werte."""

    def test_defaults_when_missing_or_empty(self) -> None:
        assert Settings._parse_customers(None) == list(DEFAULT_CUSTOMERS)
        assert Settings._parse_customers([]) == list(DEFAULT_CUSTOMERS)
        assert Settings._parse_customers(["  ", ""]) == list(DEFAULT_CUSTOMERS)

    def test_strips_and_keeps_order(self) -> None:
        assert Settings._parse_customers([" Corporate ", "Vertrieb"]) == ["Corporate", "Vertrieb"]

    def test_distinct_customers_from_db(self, service: ManualEntryService) -> None:
        for customer in ("Corporate", "Vertrieb", "Corporate", ""):
            service.add(ManualEntry(entry_date=date(2026, 7, 1), ticket="DMZ-1", hours=1.0, customer=customer))
        assert service.distinct_customers() == ["Corporate", "Vertrieb"]


class TestColorNormalization:
    """Farbeingaben werden auf RRGGBB normalisiert, ungueltiges faellt zurueck."""

    @pytest.mark.parametrize(
        ("raw", "expected"),
        [
            ("#FF0000", "FF0000"),
            ("ff0000", "FF0000"),
            ("#F00", "FF0000"),
            ("255,0,0", "FF0000"),
            ("255, 0, 0", "FF0000"),
            ("0,128,255", "0080FF"),
        ],
    )
    def test_accepted(self, raw: str, expected: str) -> None:
        assert normalize_color(raw) == expected

    @pytest.mark.parametrize("raw", ["", "rot", "#GGGGGG", "300,0,0", "12345"])
    def test_rejected_falls_back(self, raw: str) -> None:
        assert normalize_color(raw) == DEFAULT_MANUAL_COLOR


class TestExcelExport:
    """Der Export muss Spaltenkonfiguration, Kunde und Markierung umsetzen."""

    @staticmethod
    def _timesheet() -> Timesheet:
        jira = WorklogEntry(
            date=date(2026, 7, 1), ticket="DMZ-13983", summary="Agile", author="M", budget="", hours=1.0
        )
        manual = ManualEntry(
            entry_date=date(2026, 7, 1),
            ticket="DMZ-17024",
            summary="Pipelines",
            customer="Corporate",
            hours=10.0,
            entry_id=7,
        ).to_worklog(author="M")
        return Timesheet(
            developer="M",
            email="m@example.org",
            date_from=date(2026, 7, 1),
            date_to=date(2026, 7, 31),
            days=[TimesheetDay(date=date(2026, 7, 1), entries=[jira, manual])],
        )

    def test_default_export_has_customer_column(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        ExcelExporter(default_customer="Vertrieb").export(self._timesheet(), output_path=str(path))

        ws = load_workbook(path)["Stundenzettel"]
        headers = [ws.cell(row=9, column=c).value for c in range(1, 9)]
        assert headers == ["KW", "Tag", "Datum", "Ticket", "Beschreibung", "Kunde", "Aufwand (h)", "Tagessumme (h)"]
        # Jira-Eintrag faellt auf den Standard-Kunden zurueck, manueller nicht.
        assert ws.cell(row=10, column=6).value == "Vertrieb"
        assert ws.cell(row=11, column=6).value == "Corporate"

    def test_manual_row_is_marked(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        ExcelExporter(manual_color="FF0000").export(self._timesheet(), output_path=str(path))

        ws = load_workbook(path)["Stundenzettel"]
        jira_cell = ws.cell(row=10, column=4)
        manual_cell = ws.cell(row=11, column=4)
        assert manual_cell.font.bold is True
        assert manual_cell.font.color.rgb.endswith("FF0000")
        assert not jira_cell.font.bold

    def test_marking_can_be_switched_off(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        ExcelExporter(mark_manual=False).export(self._timesheet(), output_path=str(path))

        ws = load_workbook(path)["Stundenzettel"]
        assert not ws.cell(row=11, column=4).font.bold

    def test_disabled_columns_are_dropped_and_labels_applied(self, tmp_path: Path) -> None:
        columns = default_columns()
        for column in columns:
            if column.key in ("week", "weekday"):
                column.enabled = False
            if column.key == "customer":
                column.label = "Kostenstelle"

        path = tmp_path / "out.xlsx"
        ExcelExporter(columns=columns).export(self._timesheet(), output_path=str(path))

        ws = load_workbook(path)["Stundenzettel"]
        headers = [ws.cell(row=9, column=c).value for c in range(1, 7)]
        assert headers == ["Datum", "Ticket", "Beschreibung", "Kostenstelle", "Aufwand (h)", "Tagessumme (h)"]
        assert ws.cell(row=9, column=7).value is None

    def test_single_column_export_does_not_crash(self, tmp_path: Path) -> None:
        """Randfall: nur eine Spalte aktiv - Kopf und Fuss duerfen nicht kippen."""
        columns = [ExportColumn(key=c.key, label=c.label, enabled=c.key == "ticket") for c in COLUMN_DEFAULTS]
        path = tmp_path / "out.xlsx"
        ExcelExporter(columns=columns).export(self._timesheet(), output_path=str(path))

        ws = load_workbook(path)["Stundenzettel"]
        assert ws.cell(row=9, column=1).value == "Ticket"
