"""Excel Export - Stundenzettel mit KW, Wochentag und Luecken-Erkennung."""

from __future__ import annotations

import os
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from jira_timesheet.models.export_column import (
    DESCRIPTION_KEY,
    MIN_EXCEL_DESCRIPTION_WIDTH,
    ExportColumn,
    default_columns,
    excel_width,
)
from jira_timesheet.models.timesheet import Timesheet, TimesheetDay, WorklogEntry

_WEEKDAYS = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

# Gesamtbreite, die die Tabelle anstrebt (Excel-Zeicheneinheiten). Die
# Beschreibungs-Spalte fuellt auf, was die uebrigen Spalten uebrig lassen.
_TARGET_TABLE_WIDTH = 181.0

# Spalten, die rechtsbuendig ausgerichtet werden.
_RIGHT_ALIGNED = {"hours", "day_hours"}

# Was in einer Zelle stehen kann - openpyxl akzeptiert nur diese Typen.
_CellValue = str | float | int | date | None


class ExcelExporter:
    """Erzeugt eine Excel-Datei im Originalformat des Stundenzettels."""

    def __init__(
        self,
        logo_path: str = "",
        jira_host: str = "",
        hours_per_day: float = 8.0,
        show_ticket_links: bool = False,
        columns: list[ExportColumn] | None = None,
        default_customer: str = "",
        mark_manual: bool = True,
        manual_color: str = "FF0000",
    ) -> None:
        self._logo_path = logo_path
        self._jira_host = jira_host.rstrip("/")
        self._hours_per_day = hours_per_day
        self._show_ticket_links = show_ticket_links
        self._default_customer = default_customer
        self._mark_manual = mark_manual
        self._manual_color = manual_color
        source = columns if columns is not None else default_columns()
        self._columns = [c for c in source if c.enabled]

    @staticmethod
    def suggested_filename(timesheet: Timesheet) -> str:
        """Liefert einen vorgeschlagenen Dateinamen fuer den Speichern-Dialog."""
        from datetime import datetime

        now = datetime.now()
        return f"Stundenzettel_{timesheet.date_from:%Y-%m-%d}_{timesheet.date_to:%Y-%m-%d}_{now:%Y%m%d_%H%M%S}.xlsx"

    def export(
        self,
        timesheet: Timesheet,
        missing_days: list[tuple[date, str]] | None = None,
        target_hours: float = 0.0,
        output_dir: str = "",
        output_path: str = "",
    ) -> str:
        """Exportiert den Timesheet als .xlsx Datei.

        Ist ``output_path`` gesetzt, wird genau dieser Pfad verwendet (z.B. aus
        dem Speichern-Dialog). Andernfalls wird ein Name in ``output_dir`` bzw.
        auf dem Desktop erzeugt.
        """
        if output_path:
            filepath = output_path
        else:
            if not output_dir:
                output_dir = str(Path.home() / "Desktop")
            filepath = os.path.join(output_dir, self.suggested_filename(timesheet))

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Stundenzettel"

        self._setup_columns(ws)
        self._add_logo(ws)
        self._add_header(ws, timesheet, target_hours)
        self._add_table_header(ws)
        last_row = self._add_data(ws, timesheet, missing_days or [])
        self._add_footer(ws, last_row)
        self._setup_print(ws)

        wb.save(filepath)
        return str(Path(filepath).resolve())

    # --- Spalten-Hilfen ---------------------------------------------

    @property
    def _column_count(self) -> int:
        """Anzahl der aktiven Spalten."""
        return len(self._columns)

    def _setup_columns(self, ws: Worksheet) -> None:
        """Setzt die Spaltenbreiten; die Beschreibung fuellt den Rest."""
        fixed = sum(excel_width(c.key) for c in self._columns if c.key != DESCRIPTION_KEY)
        rest = max(MIN_EXCEL_DESCRIPTION_WIDTH, _TARGET_TABLE_WIDTH - fixed)

        for idx, column in enumerate(self._columns, 1):
            width = rest if column.key == DESCRIPTION_KEY else excel_width(column.key)
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _cell_value(self, entry: WorklogEntry, key: str, day: TimesheetDay, is_first: bool) -> _CellValue:
        """Liefert den Zellwert einer Spalte fuer einen Worklog-Eintrag."""
        if key == "week":
            return entry.date.isocalendar()[1] if is_first else None
        if key == "weekday":
            return _WEEKDAYS[entry.date.weekday()] if is_first else None
        if key == "date":
            return entry.date if is_first else None
        if key == "ticket":
            return entry.ticket
        if key == DESCRIPTION_KEY:
            return entry.summary
        if key == "customer":
            return entry.customer or self._default_customer
        if key == "hours":
            return entry.hours
        if key == "day_hours":
            return day.total_hours if is_first else None
        return None

    # --- Kopfbereich ------------------------------------------------

    def _add_logo(self, ws: Worksheet) -> None:
        """Fuegt das Logo in A1:B2 ein."""
        logo = self._find_logo()
        if logo and os.path.isfile(logo):
            try:
                img = XlImage(logo)
                # Seitenverhaeltnis beibehalten, max 160px breit
                ratio = img.width / img.height if img.height else 1
                img.width = 160
                img.height = int(160 / ratio)
                ws.add_image(img, "A1")
            except Exception:
                pass

    def _find_logo(self) -> str:
        """Sucht das Logo: erst Settings-Pfad, dann assets/logo.png."""
        if self._logo_path and os.path.isfile(self._logo_path):
            return self._logo_path

        here = Path(__file__).resolve().parent.parent.parent.parent
        default_logo = here / "assets" / "logo.png"
        if default_logo.is_file():
            return str(default_logo)

        return ""

    def _add_header(self, ws: Worksheet, ts: Timesheet, target_hours: float) -> None:
        """Fuegt Titel, Entwickler, Zeitraum und Gesamtstunden hinzu."""
        font_title = Font(name="Arial", size=16, bold=True)
        font_label = Font(name="Arial", size=10, bold=True)
        font_normal = Font(name="Arial", size=10)

        # Titel und Summen sitzen relativ zur Spaltenzahl, damit sie auch bei
        # abgewaehlten Spalten buendig zur Tabelle stehen.
        title_col = min(4, self._column_count)
        sum_label_col = max(1, self._column_count - 1)
        sum_value_col = max(1, self._column_count)

        if sum_label_col > title_col:
            ws.merge_cells(start_row=3, start_column=title_col, end_row=3, end_column=sum_label_col)
        cell_title = ws.cell(row=3, column=title_col)
        cell_title.value = "Stundenzettel"
        cell_title.font = font_title

        ws.cell(row=5, column=1).value = "Entwickler"
        ws.cell(row=5, column=1).font = font_label
        ws.cell(row=5, column=title_col).value = ts.developer
        ws.cell(row=5, column=title_col).font = font_normal

        ws.cell(row=6, column=1).value = "Zeitraum"
        ws.cell(row=6, column=1).font = font_label
        ws.cell(row=6, column=title_col).value = f"{ts.date_from:%d.%m.%Y} - {ts.date_to:%d.%m.%Y}"
        ws.cell(row=6, column=title_col).font = font_normal

        ws.cell(row=6, column=sum_label_col).value = "Gesamt (h)"
        ws.cell(row=6, column=sum_label_col).font = font_label
        total_cell = ws.cell(row=6, column=sum_value_col)
        total_cell.value = ts.total_hours
        total_cell.font = font_label
        total_cell.number_format = "0.00"

        if target_hours > 0:
            ws.cell(row=7, column=sum_label_col).value = "Soll (h)"
            ws.cell(row=7, column=sum_label_col).font = font_normal
            target_cell = ws.cell(row=7, column=sum_value_col)
            target_cell.value = target_hours
            target_cell.font = font_normal
            target_cell.number_format = "0.00"

    def _add_table_header(self, ws: Worksheet) -> None:
        """Fuegt die Tabellenheader-Zeile hinzu (Zeile 9)."""
        header_fill = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
        header_font = Font(name="Arial", size=10)
        header_align = Alignment(horizontal="left", vertical="center")

        for col_idx, column in enumerate(self._columns, 1):
            cell = ws.cell(row=9, column=col_idx)
            cell.value = column.label
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    # --- Daten ------------------------------------------------------

    def _add_data(
        self,
        ws: Worksheet,
        ts: Timesheet,
        missing_days: list[tuple[date, str]],
    ) -> int:
        """Fuegt die Datenzeilen hinzu. Gibt die letzte Zeilennummer zurueck."""
        day_map = {day.date: day for day in ts.days}
        gap_map = dict(missing_days)
        all_dates = sorted(set(day_map.keys()) | set(gap_map.keys()))

        current_row = 10
        for d in all_dates:
            if d in gap_map and d not in day_map:
                current_row = self._write_gap_row(ws, d, gap_map[d], current_row)
                continue
            if d not in day_map:
                continue
            current_row = self._write_day(ws, day_map[d], current_row)

        return current_row - 1

    def _write_gap_row(self, ws: Worksheet, d: date, reason: str, row: int) -> int:
        """Schreibt eine Luecken-/Feiertagszeile. Gibt die naechste Zeile zurueck."""
        font_dim = Font(name="Arial", size=10, color="999999")
        font_red = Font(name="Arial", size=10, color="CC0000")
        align_left = Alignment(horizontal="left", vertical="center")
        border_day_top = Border(top=Side(style="medium", color="000000"))

        ws.row_dimensions[row].height = 20.1
        is_holiday = "—" not in reason
        use_font = font_dim if is_holiday else font_red

        values: dict[str, _CellValue] = {
            "week": d.isocalendar()[1],
            "weekday": _WEEKDAYS[d.weekday()],
            "date": d,
            DESCRIPTION_KEY: reason,
            "day_hours": 0,
        }
        for col_idx, column in enumerate(self._columns, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = values.get(column.key)
            cell.font = use_font
            cell.alignment = align_left
            cell.border = border_day_top
            if column.key == "date":
                cell.number_format = "DD.MM.YYYY"
            elif column.key == "day_hours":
                cell.number_format = "0.00"

        return row + 1

    def _write_day(self, ws: Worksheet, day: TimesheetDay, row: int) -> int:
        """Schreibt alle Eintraege eines Tages. Gibt die naechste Zeile zurueck."""
        font_normal = Font(name="Arial", size=10)
        font_manual = Font(name="Arial", size=10, bold=True, color=self._manual_color)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        border_day_top = Border(top=Side(style="medium", color="000000"))

        for i, entry in enumerate(day.entries):
            is_first = i == 0
            ws.row_dimensions[row].height = 20.1
            mark = entry.manual and self._mark_manual

            for col_idx, column in enumerate(self._columns, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = self._cell_value(entry, column.key, day, is_first)
                cell.font = font_manual if mark else font_normal
                cell.alignment = align_right if column.key in _RIGHT_ALIGNED else align_left

                if column.key == "date":
                    cell.number_format = "DD.MM.YYYY"
                elif column.key in _RIGHT_ALIGNED:
                    cell.number_format = "0.00"
                elif column.key == "ticket":
                    self._apply_ticket_link(cell, entry, mark)

                if is_first:
                    cell.border = border_day_top

            row += 1

        return row

    def _apply_ticket_link(self, cell: Cell | MergedCell, entry: WorklogEntry, mark: bool) -> None:
        """Haengt den Jira-Link an die Ticket-Zelle, falls aktiviert."""
        if not self._show_ticket_links or not self._jira_host or not entry.ticket:
            return
        cell.hyperlink = f"{self._jira_host}/browse/{entry.ticket}"
        # Manuelle Markierung schlaegt die Link-Farbe, sonst geht sie unter.
        color = self._manual_color if mark else "0563C1"
        cell.font = Font(name="Arial", size=10, bold=mark, color=color, underline="single")

    # --- Fuss und Druck ---------------------------------------------

    def _add_footer(self, ws: Worksheet, last_row: int) -> None:
        """Fuegt die Unterschriftszeile hinzu."""
        footer_row = last_row + 3
        font_normal = Font(name="Arial", size=10)
        border_top = Border(top=Side(style="thin", color="000000"))

        label_col = min(4, self._column_count)
        text_col = min(5, self._column_count)

        ws.cell(row=footer_row, column=label_col).value = "bestätigt am:"
        ws.cell(row=footer_row, column=label_col).font = font_normal

        if text_col > label_col:
            ws.cell(row=footer_row, column=text_col).value = "Projektleiter (Blockschrift, Unterschrift)"
            ws.cell(row=footer_row, column=text_col).font = font_normal

        for col in range(label_col, self._column_count + 1):
            ws.cell(row=footer_row, column=col).border = border_top

    def _setup_print(self, ws: Worksheet) -> None:
        """Setzt Druckeinstellungen: A4 Portrait, Fit to Page."""
        try:
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            pass
